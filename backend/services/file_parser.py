"""文件解析服务：把上传的多种格式文件解析为纯文本，供对话上下文注入。"""
import csv
import io


def _parse_text(content: bytes) -> str:
    """txt / md：直接按 UTF-8 解码。"""
    return content.decode("utf-8", errors="replace")


def _parse_csv(content: bytes) -> str:
    """csv：解析后用空格对齐成表格文本。"""
    decoded = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(decoded))
    rows = list(reader)
    if not rows:
        return ""
    # 计算每列最大宽度，用于对齐
    col_widths = [
        max(len(row[i]) if i < len(row) else 0 for row in rows)
        for i in range(len(rows[0]))
    ]
    lines = []
    for row in rows:
        padded = []
        for i, cell in enumerate(row):
            if i < len(col_widths):
                padded.append(cell.ljust(col_widths[i]))
            else:
                padded.append(cell)
        lines.append("  ".join(padded))
    return "\n".join(lines)


def _parse_pdf(content: bytes) -> str:
    """pdf：逐页提取文本。"""
    from pypdf import PdfReader
    reader = PdfReader(io.BytesIO(content))
    pages = [page.extract_text() for page in reader.pages]
    return "\n\n".join(t for t in pages if t)


def _parse_docx(content: bytes) -> str:
    """docx：提取非空段落文本。"""
    from docx import Document
    doc = Document(io.BytesIO(content))
    paras = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n".join(paras)


def _parse_xlsx(content: bytes) -> str:
    """xlsx：逐 sheet 逐行导出为制表符分隔文本。"""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    lines = []
    for idx, sheet in enumerate(wb.worksheets):
        if idx > 0:
            lines.append("")  # 空行分隔工作表
        lines.append(f"=== {sheet.title} ===")
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            lines.append("\t".join(cells))
    return "\n".join(lines)


# 扩展名 -> 解析函数注册表
PARSERS = {
    ".txt": _parse_text,
    ".md": _parse_text,
    ".csv": _parse_csv,
    ".pdf": _parse_pdf,
    ".docx": _parse_docx,
    ".xlsx": _parse_xlsx,
}


def parse_content(content: bytes, ext: str) -> str:
    """按扩展名分发到对应的解析器，未知类型返回空串。"""
    parser = PARSERS.get(ext)
    return parser(content) if parser else ""
