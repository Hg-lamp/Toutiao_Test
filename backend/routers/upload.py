import os
import csv
import io
from pathlib import Path

from fastapi import APIRouter, File, UploadFile, HTTPException
from pydantic import BaseModel

router = APIRouter(prefix="/api", tags=["upload"])

# 支持的文件类型
ALLOWED_EXTENSIONS = {
    ".txt", ".md", ".csv",
    ".pdf", ".docx", ".xlsx",
}

# 文件大小限制：5MB
MAX_FILE_SIZE = 5 * 1024 * 1024


class UploadResponse(BaseModel):
    filename: str
    text: str
    size: int


@router.post("/upload", response_model=UploadResponse)
async def upload_file(file: UploadFile = File(...)):
    """上传文件，解析文本内容后返回，不存盘。"""
    # 1. 校验文件类型
    ext = Path(file.filename or "").suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"不支持的文件类型: {ext}，仅支持 {', '.join(ALLOWED_EXTENSIONS)}"
        )

    # 2. 读取文件内容（限制大小）
    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(
            status_code=400,
            detail=f"文件过大（{len(content) / 1024 / 1024:.1f}MB），最大支持 5MB"
        )

    # 3. 根据文件类型解析文本
    try:
        text = _parse_content(content, ext)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"文件解析失败: {str(e)}")

    # 4. 限制文本长度（防止前端炸掉）
    MAX_CHARS = 100000  # 约 10 万字
    if len(text) > MAX_CHARS:
        text = text[:MAX_CHARS] + "\n\n...（文件过长，已截断）"

    return UploadResponse(
        filename=file.filename or "unknown",
        text=text,
        size=len(content),
    )


def _parse_content(content: bytes, ext: str) -> str:
    """根据文件扩展名解析文本内容。"""
    if ext in (".txt", ".md"):
        return content.decode("utf-8", errors="replace")

    if ext == ".csv":
        # 用 csv 模块解析，转成表格文本
        decoded = content.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(decoded))
        rows = list(reader)
        if not rows:
            return ""
        # 计算每列最大宽度对齐
        col_widths = [
            max(len(row[i]) if i < len(row) else 0 for row in rows)
            for i in range(len(rows[0]))
        ]
        lines = []
        for row in rows:
            padded = []
            for i, cell in enumerate(row):
                if i < len(col_widths):
                    padded.append(cell.ljust(col_widths[i]))
                else:
                    padded.append(cell)
            lines.append("  ".join(padded))
        return "\n".join(lines)

    if ext == ".pdf":
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(content))
        pages = []
        for page in reader.pages:
            t = page.extract_text()
            if t:
                pages.append(t)
        return "\n\n".join(pages)

    if ext == ".docx":
        from docx import Document
        doc = Document(io.BytesIO(content))
        paras = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n".join(paras)

    if ext == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        lines = []
        for sheet in wb.worksheets:
            if wb.worksheets.index(sheet) > 0:
                lines.append("")  # 空行分隔工作表
            lines.append(f"=== {sheet.title} ===")
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                lines.append("\t".join(cells))
        return "\n".join(lines)

    return ""